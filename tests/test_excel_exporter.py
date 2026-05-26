from openpyxl import load_workbook
import pandas as pd

from vpa_structure_recognizer.excel_exporter import REQUIRED_SHEETS, export_excel_report


def test_export_excel_report_writes_required_sheets(tmp_path):
    output = tmp_path / "vpa_structure_report_20240131.xlsx"
    market = pd.DataFrame([{"date": "2024-01-31", "market_score": 80, "final_rating": "A"}])
    sectors = pd.DataFrame([{"date": "2024-01-31", "scope_id": "BK001", "final_rating": "A"}])
    stocks = pd.DataFrame(
        [
            {
                "date": "2024-01-31",
                "scope_id": "000001.SZ",
                "final_state": "HEALTHY_UPTREND",
                "final_rating": "A",
            }
        ]
    )
    labels = pd.DataFrame(
        [{"date": "2024-01-31", "scope_id": "000001.SZ", "window_n": 20, "raw_label": "NORMAL_UP_CONFIRM"}]
    )
    validation = pd.DataFrame(
        [{"date": "2024-01-31", "scope_id": "000001.SZ", "future_ret_5d": 0.05}]
    )

    export_excel_report(output, market, sectors, stocks, labels, validation)

    workbook = load_workbook(output)
    assert workbook.sheetnames == REQUIRED_SHEETS
    assert workbook["个股结构总表"]["A1"].value == "date"
    assert workbook["个股多窗口标签明细"]["D1"].value == "window_n"
